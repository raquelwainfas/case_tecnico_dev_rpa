from openpyxl import Workbook, load_workbook
import os

class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        
    def create_excel(self, columns=[]):
        """Cria um arquivo Excel com as colunas especificadas"""

        if os.path.exists(self.filename):
            print(f"Arquivo {self.filename} já existe. Não será sobrescrito.")
            return

        wb = Workbook()
        ws = wb.active
        
        # Preenche as colunas (cabeçalho)
        for col_num, column_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)
        
        wb.save(self.filename)
        
    def append_row(self, data_list):
        """Anexa uma linha ao arquivo Excel existente"""
        if not os.path.exists(self.filename):
            raise FileNotFoundError(f"Arquivo {self.filename} não encontrado")
            
        wb = load_workbook(self.filename)
        ws = wb.active
        
        # Encontra a próxima linha vazia
        next_row = ws.max_row + 1
        
        # Adiciona os dados na nova linha
        for col_num, value in enumerate(data_list, 1):
            ws.cell(row=next_row, column=col_num, value=value)
        
        wb.save(self.filename)